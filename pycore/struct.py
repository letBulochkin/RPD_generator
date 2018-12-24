import openpyxl
from enum import Enum
import xml.etree.ElementTree as etree
from . import crawler

class study_plan(object):
    """
    Study plan class.

    TODO: реализовать все алгоритмы поиска отдельно.

    Errors:
    xml.etree.ElementTree.ParseError: mismatched tag:
    Ненайденный элемент

    """

    def __init__(self, book, rules):
        '''
        Инициализация полей класса
        Class attributes initialization

        TODO: oh we need mapper
        '''

        self.BOOK = openpyxl.load_workbook(book)
        self.DATA = etree.parse(rules).getroot()

        sp = self.DATA.find('studyplan')
        if sp is None:
            raise ValueError("Necessary tag not found: ", 'studyplan')

        sheet = self.BOOK[sp.find('sheet').text]
        
        self.MINISTRY = sheet[sp.find('ministry').text].value
        self.UNIVERCITY = sheet[sp.find('univercity').text].value.rsplit('\n', 1)[0]
        self.INSTITUTE = sheet[sp.find('institute').text].value.rsplit('\n', 1)[1]
        self.RECTOR = sheet[sp.find('rector').text].value
        self.FIELD_OF_KNOW = sheet[sp.find('field').text].value.rsplit('Направление ', 1)[1]
        self.PROFILE = sheet[sp.find('profile').text].value
        self.CATHEDRA = sheet[sp.find('cath').text].value
        self.QUALI_LEVEL = sheet[sp.find('quali_level').text].value.rsplit('Квалификация: ', 1)[1]
        self.EDU_PROG = sheet[sp.find('edu_prog').text].value.rsplit('Программа подготовки: ', 1)[1]
        self.EDU_FORMAT = sheet[sp.find('edu_format').text].value.rsplit('Форма обучения: ', 1)[1]
        self.EDU_TIME = sheet[sp.find('edu_time').text].value.rsplit('Срок обучения: ', 1)[1]  # поправить, убрать приписку г
        self.START_YEAR = sheet[sp.find('start_year').text].value

        start = sp.find('practice_types').find('start').text
        stop = sp.find('practice_types').find('stop').text

        self.PRACTICE_TYPES = [[i[2]] for i in crawler.range_search(
            sheet, start, stop, '', mode = "notmatch")]

    def data_parse(self):
        pass

    def list_avail_disciplines(self):
        '''
        Возвращает список изучаемых на кафедре дисцпилин
        Returns list of available disciplines
        '''

        disciplines = []

        ad = self.DATA.find('studyplan').find('avail_disciplines')
        if ad is None:
            raise ValueError("Necessary tag not found: ", 'avail_disciplines')
        
        sheet = self.BOOK[ad.find('sheet').text]

        start = ad.find('start').text
        stop = ad.find('stop').text
        code_col = int(ad.find('code_col').text)
        descrp_col = int(ad.find('descrp_col').text)
        for i in crawler.range_search(sheet, start, stop, self.CATHEDRA):
            disciplines.append([sheet.cell(row = i[0], column = code_col).value, 
                sheet.cell(row = i[0], column = descrp_col).value]) 
        
        return disciplines

class discipline(object):

    def __init__(self, study_plan, index):
        
        self.STUDY_PLAN = study_plan
        self.INDEX = index
        self.NAME = ''.join(i[1] for i in self.STUDY_PLAN.list_avail_disciplines() if i[0] == index)
        
        self.TAG = self.STUDY_PLAN.DATA.find('discipline')
        if self.TAG is None:
            raise ValueError("Necessary tag not found: ", 'discipline')

        if index.rsplit('.')[1] == 'Б': 
            self.PART = True  # базовая
        elif index.rsplit('.')[1] == 'В':
            self.PART = False  # вариативная 

        if index.rsplit('.')[2] == 'ДВ':
            self.OBLIGATION = False  # по выбору
        else:
            self.OBLIGATION = True

        self.COMPETENCIES = self.__get_competencies__()
        self.STUDY_HOURS = self.__get_hours__()
        self.SEMESTERS = self.__get_semesters__()
        self.LABS = []
        self.PRACT = []

    def __get_competencies__(self):
        '''
        Возвращает список компетенций, изучаемых дисциплиной, и их описаний
        Returns dictionary of competencies and theirs descriptions
        '''
        
        competencies = []
        
        com = self.TAG.find('competencies')

        sheet = self.STUDY_PLAN.BOOK[com.find('indexes').find('sheet').text]
        
        r = crawler.range_search(
            sheet, 
            com.find('indexes').find('start').text, 
            com.find('indexes').find('stop').text, 
            self.INDEX)[0][0]

        for i in sheet.cell(row = r, column = 6).value.rsplit('; '):
            q = {}
            q['code'] = i
            sheet = self.STUDY_PLAN.BOOK[com.find('descriptions').find('sheet').text]
            cont = sheet.cell(
                row = crawler.range_search(
                    sheet, 
                    com.find('descriptions').find('start').text, 
                    com.find('descriptions').find('stop').text, 
                    i)[0][0], 
                column = int(com.find('descriptions').find('descrp_col').text)).value
            q['descrp'] = cont
            q['part'] = None
            q['to_know'] = None
            q['to_can'] = None
            q['to_be_able'] = None
            competencies.append(q)

        return competencies

    def __get_semesters__(self):
        """
        Нет, это полная хуйня.
        Но в принципе работает. 
        """

        semesters = []  # список семестров, в которые читается дисциплина

        sems = self.TAG.find('semesters')

        sheet = self.STUDY_PLAN.BOOK[sems.find('sheet').text]

        dicp_cell = crawler.range_search(
            sheet, 
            sems.find('start').text, 
            sems.find('stop').text, 
            self.INDEX)  # поиск ячейки с дисциплиной 

        # задаем диапазон для поиска значений по найденной ячейке
        search_cell_start = crawler.coord_to_letter(
            dicp_cell[0][0], 
            dicp_cell[0][1] + int(sems.find('sem_search').find('start_sdv').text))
        search_cell_stop = crawler.coord_to_letter(
            dicp_cell[0][0], 
            dicp_cell[0][1] + int(sems.find('sem_search').find('stop_sdv').text))

        # по тем ячейкам, где было найдено значение, поднимаемся наверх и смотрим номер семестра
        for i in crawler.range_search(sheet, search_cell_start, search_cell_stop, None, "notmatch"):
            q = []
            sem_no = int(sheet.cell(
                row = int(sems.find('sem_search').find('sem_no_row').text), 
                column = i[1]
            ).value.rsplit('. ', 1)[1])  # получаем номер семестра
            q.extend([sem_no, semester(), False, False, False, False])  # абсолютно отвратительно
            search_st = crawler.coord_to_letter(
                dicp_cell[0][0], 
                dicp_cell[0][1] + int(sems.find('control_search').find('start_sdv').text))
            search_fin = crawler.coord_to_letter(
                dicp_cell[0][0], 
                dicp_cell[0][1] + int(sems.find('control_search').find('stop_sdv').text))
            for k in crawler.range_search(sheet, search_st, search_fin, str(sem_no), "in"):
                if sheet.cell(row = int(sems.find('control_search').find('control_row').text), column = k[1]).value == 'Экза мен':
                    q[2] = True
                elif sheet.cell(row = int(sems.find('control_search').find('control_row').text), column = k[1]).value == 'Зачет':
                    q[3] = True
                elif sheet.cell(row = int(sems.find('control_search').find('control_row').text), column = k[1]).value == 'Зачет c оц.':
                    q[4] = True
                elif sheet.cell(row = int(sems.find('control_search').find('control_row').text), column = k[1]).value == 'КР':
                    q[5] = True
            semesters.append(q)

        return semesters

    def __get_hours__(self):

        sem = self.__get_semesters__()

        hrs = self.TAG.find('hours')

        sheet = self.STUDY_PLAN.BOOK[hrs.find('sheet').text]
        res = []

        for i in range(len(sem)):
            h = {}
            dicp_cell = crawler.range_search(
                sheet, 
                hrs.find('dicp_cell').find('start').text, 
                hrs.find('dicp_cell').find('stop').text, 
                self.INDEX)
            sem_cell = crawler.range_search(
                sheet, 
                hrs.find('sem_cell').find('start').text, 
                hrs.find('sem_cell').find('stop').text, 
                'Сем. ' + str(sem[i][0]))
            h['zach_ed'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1]).value)
            h['total'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1] + 1).value)
            h['lect'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1] + 2).value)
            h['lab'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1] + 3).value)
            h['pract'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1] + 4).value)
            h['sam'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1] + 5).value)
            h['krpa'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1] + 6).value)
            h['control'] = crawler.int_eater(sheet.cell(row = dicp_cell[0][0], column = sem_cell[0][1] + 7).value)
            res.append([sem[i][0], h])

        return res

class semester(object):

    def __init__(self):

        self.STUDY_TYPES = Enum('Study_Types', 'lect lab pract sam')

        # self.EXAM = None
        # self.ZACHET = None
        # self.CONTROL_FORMS = []
        
        self.MODULES = []
        
    def add_module(self, num, description):

        res = {}

        res['num'] = int(num)
        res['descr'] = description
        res['lect'] = None
        res['lab'] = None
        res['pract'] = None
        res['sam'] = None

        self.MODULES.append(res)

    def add_study(self, module, type, hours):

        for i in self.MODULES:
            if i['num'] == module:
                t = self.STUDY_TYPES(type).name
                i[t] = hours

    """
    def add_study(self, module, type, num, hours, description):

        for i in self.MODULES:
            if i['num'] == module:
                t = self.STUDY_TYPES(type).name
                i[t] = [num, description, hours]
    """
    """
    def set_sam_hours(self, module, hours):

        for i in self.MODULES:
            if i['num'] == module:
                i['sam'] = hours
    """
